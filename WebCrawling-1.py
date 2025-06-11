import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

Headers = {'User-Agent' : 'Mozilla/5.0 (Window NT 10.0 : Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) chrome / 128.0.0.0 Sarari/537.36'}
Url = "https://www.melon.com/new/index.htm"
Req = requests.get(Url,headers = Headers)
Html = Req.text
Soup = BeautifulSoup(Html,'html.parser')
NewList50 = Soup.select('td > div')

for Memo in NewList50:
    Rank = Memo.select_one('span.rank')

    if Rank is not None:
        print(Rank.text, end = '위. ')

    Title = Memo.select_one('div > div.ellipsis.rank01 > span > a')

    if Title is not None:
        Artist = Memo.select_one('div.ellipsis.rank02 > a').text
        Album = Memo.select_one('div.ellipsis.rank03 > a')
        print(Artist, ',', Title.text, ',', Album)

        wb = Workbook()
        ws = wb.active
        ws.title = "멜론 최신곡 50"
        ws.merge_cells('A1:E1')
        ws['A1'] = '멜론최신곡50'
        column = ['순위','아티스티','노래제목','앨범명']
        ws.append(column)
        row = [Rank,Artist,Title,Album]
        prefix1 = 'A'
        prefix2 = 'B'
        prefix3 = 'C'
        prefix4 = 'D'

        num1 = 3
        while num1 < 53 :
            str1 = f'{prefix1}{num1}'
            str2 = f'{prefix2}{num1}'
            str3 = f'{prefix3}{num1}'
            str4 = f'{prefix4}{num1}'
            ws[str1] = Rank
            ws[str2] = Artist
            ws[str3] = Title
            ws[str4] = Album
            print(ws[str1],ws[str2],ws[str3],ws[str4])
            num1 = num1 + 1
